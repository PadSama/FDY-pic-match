import openpyxl
import base64
import io
from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import ZipFile
import xml.etree.ElementTree as ET
import os

app = Flask(__name__)
CORS(app)  # 保持跨域支持

@app.route('/api/process_excel', methods=['POST'])
def process_excel():
    print("=== 收到文件上传请求 ===")
    try:
        # 校验是否上传了文件
        if 'file' not in request.files:
            return jsonify({'error': '未上传Excel文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '上传的文件为空'}), 400
        
        if not file.filename.endswith('.xlsx'):
            return jsonify({'error': '仅支持.xlsx格式的Excel文件'}), 400
        
        try:
            wb = openpyxl.load_workbook(file)
        except InvalidFileException:
            return jsonify({'error': '文件格式错误，不是有效的Excel文件'}), 400
        
        sheet = wb.active
        
        if sheet.max_row < 2:
            return jsonify({'error': 'Excel文件无有效数据'}), 400
        
        results = []
        total_rows = 0
        success_rows = 0

        # 新增：计算每行的垂直范围（基于行高）及行高验证日志
        row_heights = []
        for row_idx in range(1, sheet.max_row + 1):
            height = sheet.row_dimensions[row_idx].height or 15  # 默认15磅（Excel默认行高）
            row_heights.append(height)
        
        row_y_ranges = []
        current_y = 0.0
        for idx, height in enumerate(row_heights, start=1):
            row_y_ranges.append({
                'row': idx,
                'y_start': current_y,
                'y_end': current_y + height
            })
            current_y += height
        
        print("=== 行高范围 ===")
        for row_range in row_y_ranges:
            print(f"行{row_range['row']}: y_start={row_range['y_start']:.2f}, y_end={row_range['y_end']:.2f}")

        # 新增：通过ZIP提取浮动图片并匹配垂直范围
        image_cache = {}  # 键：行号，值：图片Base64
        with ZipFile(file.stream, 'r') as zip_ref:
            # 提取绘图文件时去重（关键优化）
            drawing_files = list(set([f for f in zip_ref.namelist() if f.startswith('xl/drawings/')]))
            drawing_files.sort()  # 确保顺序稳定
            
            # 解析绘图文件时跳过已处理文件（关键优化）
            processed_drawings = set()
            for drawing_path in drawing_files:
                if drawing_path in processed_drawings:
                    continue
                processed_drawings.add(drawing_path)
                # 移除多余的外层 try 块（原语法错误位置）
                with zip_ref.open(drawing_path) as drawing_file:
                    # 关键修改：保留内层 XML 解析异常处理
                    try:
                        drawing_xml = ET.parse(drawing_file).getroot()
                    except ET.ParseError as e:
                        print(f"警告：绘图文件 {drawing_path} XML解析失败（可能为空或格式错误）: {str(e)}")
                        continue  # 跳过异常文件
                    ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                          'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                          'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                        
                    # 提取所有类型的图片锚点（整合优化代码）
                    anchors = []
                    anchors += drawing_xml.findall('.//xdr:oneCellAnchor', namespaces=ns)
                    anchors += drawing_xml.findall('.//xdr:twoCellAnchor', namespaces=ns)
                    anchors += drawing_xml.findall('.//xdr:absoluteAnchor', namespaces=ns)
                        
                    for anchor in anchors:
                            anchor_type = anchor.tag.split('}')[-1]
                            from_pos = anchor.find('xdr:from', namespaces=ns)
                            pos_node = anchor.find('xdr:pos', namespaces=ns)  # 处理absoluteAnchor的绝对位置

                            # 提取y坐标（针对twoCellAnchor调整逻辑）
                            y_points = None
                            if anchor_type == 'twoCellAnchor' and from_pos is not None:
                                # 提取行号（0-based）和行内偏移量（EMU）
                                row_node = from_pos.find('xdr:row', namespaces=ns)
                                row_off_node = from_pos.find('xdr:rowOff', namespaces=ns)
                                if row_node is not None and row_off_node is not None:
                                    try:
                                        # 转换为1-based行号（与row_y_ranges对应）
                                        excel_row = int(row_node.text) + 1  # 0-based → 1-based
                                        row_off_emu = int(row_off_node.text)
                                        # 从row_y_ranges中获取该行的起始y坐标
                                        if 1 <= excel_row <= len(row_y_ranges):
                                            row_start_y = row_y_ranges[excel_row - 1]['y_start']
                                            # 计算绝对y坐标（行起始y + 偏移量转磅）
                                            y_points = row_start_y + (row_off_emu / 12700)
                                    except (ValueError, IndexError):
                                        print(f"警告：twoCellAnchor行号或偏移量解析失败，跳过")
                                        continue
                            else:
                                # 其他锚点类型保持原逻辑（oneCellAnchor/absoluteAnchor）
                                y_node = None
                                if from_pos is not None:
                                    y_node = from_pos.find('xdr:y', namespaces=ns)
                                elif pos_node is not None:
                                    y_node = pos_node.find('xdr:y', namespaces=ns)
                                if y_node is not None:
                                    try:
                                        y_emu = int(y_node.text)
                                        y_points = y_emu / 12700  # EMU转磅
                                    except ValueError:
                                        print(f"警告：y坐标值无效，跳过")
                                        continue

                            if y_points is None:
                                print(f"警告：锚点类型{anchor_type}无有效y坐标，跳过")
                                continue
                            print(f"处理锚点类型{anchor_type}，y坐标（磅）={y_points:.2f}")

                            # 修正注释符号
                            # 原有逻辑（已修正ns定义位置）
                            
                            # 匹配行号（允许±0.5磅误差）
                            # 预计算行高范围的y_start列表（按顺序排列）
                            y_starts = [row_range['y_start'] for row_range in row_y_ranges]
                            y_ends = [row_range['y_end'] for row_range in row_y_ranges]
                            
                            # 匹配行号时使用二分查找（关键优化）
                            import bisect  # 顶部新增导入
                            
                            # 原匹配逻辑替换为：
                            matched_row = None
                            tolerance = 0.5
                            # 找到第一个y_start大于(y_points - tolerance)的索引
                            index = bisect.bisect_left(y_starts, y_points - tolerance)
                            # 检查前一个索引（可能包含目标y值）
                            if index > 0:
                                candidate = index - 1
                                if (y_starts[candidate] - tolerance <= y_points <= y_ends[candidate] + tolerance):
                                    matched_row = candidate + 1  # 转换为1-based行号
                            # 检查当前索引
                            if index < len(y_starts):
                                if (y_starts[index] - tolerance <= y_points <= y_ends[index] + tolerance):
                                    matched_row = index + 1  # 转换为1-based行号

                            # 检查 matched_row 是否有效（避免未定义）
                            if matched_row is None or matched_row == 1:
                                print(f"未找到匹配行（当前y={y_points:.2f}）")
                                continue  # 跳过无效行

                            # 新增：查找a:blip节点并校验
                            blip = anchor.find('.//a:blip', namespaces=ns)
                            if blip is None:
                                print("警告：未找到a:blip节点，跳过")
                                continue

                            #// 原有逻辑（已修正ns定义位置）
                            #// 定义命名空间时补充关系命名空间（关键修改）
                            ns = {
                                'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                                'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                                'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                                'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'  # 新增关系命名空间
                            }

                            #// 提取图片媒体路径（关键修改）
                            rel_id = blip.get(f"{{{ns['r']}}}embed")
                            if not rel_id:
                                print("警告：a:blip节点无r:embed属性，跳过")
                                continue

                            rels_path = f"xl/drawings/_rels/{os.path.basename(drawing_path)}.rels"
                            try:
                                with zip_ref.open(rels_path) as rels_file:
                                    rels_xml = ET.parse(rels_file).getroot()
                                    for rel in rels_xml.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                                        if rel.attrib['Id'] == rel_id:
                                            # 移除Target中的"../"，直接指向xl/media目录
                                            target = rel.attrib['Target'].replace('../', '')  # 关键修改：将"../media/image117.jpeg"转为"media/image117.jpeg"
                                            media_path = f"xl/{target}"  # 拼接后路径为"xl/media/image117.jpeg"
                                            break
                            except KeyError:
                                print(f"警告：未找到关系文件{rels_path}，跳过")
                                continue

                            # 读取并缓存图片（保持原逻辑）
                            try:
                                with zip_ref.open(media_path) as img_file:
                                    img_bytes = io.BytesIO(img_file.read())
                                    image_data = base64.b64encode(img_bytes.getvalue()).decode('utf-8')
                                    image_cache[matched_row] = image_data
                                    print(f"成功缓存行{matched_row}的图片")
                            except KeyError:
                                print(f"警告：未找到图片文件{media_path}，跳过")

        # 标题行检测仅执行一次（关键优化）
        possible_header_rows = [1, 2, 3]
        header_row_found = None
        for candidate_row in possible_header_rows:
            if candidate_row > sheet.max_row:
                continue
            candidate_cells = [cell.value for cell in sheet[candidate_row]]
            cleaned_cells = [str(cell).strip() if cell is not None else '' for cell in candidate_cells]
            if '姓名' in cleaned_cells:
                header_row_found = candidate_row
                break
        if header_row_found is None:
            print("错误：1-3行中未找到包含'姓名'的标题行")
            raise ValueError("Excel文件中1-3行均未找到'姓名'列")

        # 数据行循环（关键优化：移除循环内的标题行检测）
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            total_rows += 1
            if row_idx < header_row_found:  # 跳过标题行前的冗余行（仅判断一次）
                continue
            if row_idx == header_row_found:  # 处理标题行（仅执行一次）
                column_mapping = {
                    'name_col': None,
                    'image_cols': []
                }
                # 新增调试日志：输出实际找到的标题行信息
                print(f"=== 找到标题行（第{header_row_found}行） ===")
                raw_columns = [str(cell_value) if cell_value is not None else '' for cell_value in row]
                cleaned_columns = [str(cell_value).strip() if cell_value is not None else '' for cell_value in row]
                print(f"原始列值: {raw_columns}")
                print(f"清理后列值: {cleaned_columns}")
                
                for col_idx, cell_value in enumerate(row):
                    cell_value_clean = str(cell_value).strip() if cell_value is not None else ''
                    print(f"检查第{col_idx+1}列：原始值='{cell_value}'，清理后='{cell_value_clean}'")
                    
                    if cell_value_clean == '姓名':  # 匹配去空格后的'姓名'
                        if column_mapping['name_col'] is not None:
                            print(f"错误：第{row_idx}行发现多个'姓名'列")
                            raise ValueError("Excel文件中存在多个'姓名'列")
                        column_mapping['name_col'] = col_idx
                    else:
                        # 强化图片列匹配逻辑（新增关键词并统一小写）
                        col_name = str(cell_value).strip().lower()  # 统一转为小写匹配
                        if any(keyword in col_name for keyword in ['图', '图片', '示意图', '照片', 'img', 'photo']) \
                           or 'image' in col_name:
                            column_mapping['image_cols'].append(col_idx)
                
                # 验证列名
                if column_mapping['name_col'] is None:
                    print(f"错误：标题行（第{header_row_found}行）未找到'姓名'列")
                    raise ValueError("Excel文件中缺少'姓名'列")
                if not column_mapping['image_cols']:
                    print(f"错误：标题行（第{header_row_found}行）未找到包含'图'或'image'的图片列")
                    raise ValueError("Excel文件中缺少图片列")
                continue
            
            # 处理数据行（从标题行的下一行开始）
            if column_mapping['name_col'] >= len(row):
                print(f"第{row_idx}行：列数不足，跳过")
                continue
            
            # 提取姓名并去除前后空格（关键修改）
            name_cell = row[column_mapping['name_col']]
            name = str(name_cell).strip() if name_cell is not None else ''  # 处理空格和空值
            if not name:
                print(f"第{row_idx}行：姓名为空或仅包含空格，跳过")
                continue
            
            # 直接通过行号获取缓存图片（修改点）
            image_data = image_cache.get(row_idx)
            
            if image_data:
                results.append({
                    'name': name,  # 已处理空格的姓名
                    'image': f"data:image/png;base64,{image_data}"
                })
                success_rows += 1
                print(f"第{row_idx}行：图片匹配成功")
            else:
                print(f"第{row_idx}行：未找到匹配的图片")

        
        print(f"处理完成：总{total_rows}行，成功{success_rows}行，返回数据长度：{len(results)}")
        return jsonify(results)
    
    except Exception as e:
        print(f"处理异常：{str(e)}")
        return jsonify({'error': f'服务器内部错误：{str(e)}'}), 500

if __name__ == '__main__':
    app.run(port=5000)